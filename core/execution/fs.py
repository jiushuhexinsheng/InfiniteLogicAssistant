# -*- coding: utf-8 -*-
"""文件系统 — 读写所有通用格式（text/json/yaml/toml/csv/xlsx/sqlite/ini/env/md）

按扩展名分发 reader/writer；未知扩展名按纯文本处理。
"""
import csv
import json
import sqlite3
from pathlib import Path
from typing import Any

import yaml


def _ext(path: Path) -> str:
    return path.suffix.lower().lstrip(".")


# ── 各格式读写器 ──

def _read_json(p): return json.loads(p.read_text(encoding="utf-8"))
def _write_json(p, data): p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def _read_yaml(p): return yaml.safe_load(p.read_text(encoding="utf-8"))
def _write_yaml(p, data): p.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")

def _read_toml(p):
    import tomllib
    return tomllib.loads(p.read_text(encoding="utf-8"))
def _write_toml(p, data):
    import tomli_w
    p.write_text(tomli_w.dumps(data), encoding="utf-8")

def _read_csv(p):
    with p.open("r", encoding="utf-8", newline="") as f:
        return [row for row in csv.reader(f)]
def _write_csv(p, data):
    with p.open("w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows(data)

def _read_ini(p):
    import configparser
    cp = configparser.ConfigParser()
    cp.read(p, encoding="utf-8")
    return {s: dict(cp.items(s)) for s in cp.sections()}
def _write_ini(p, data):
    import configparser
    cp = configparser.ConfigParser()
    for section, items in data.items():
        cp[section] = {k: str(v) for k, v in items.items()}
    with p.open("w", encoding="utf-8") as f:
        cp.write(f)

def _read_env(p):
    out = {}
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        out[k.strip()] = v.strip()
    return out
def _write_env(p, data):
    p.write_text("\n".join(f"{k}={v}" for k, v in data.items()) + "\n", encoding="utf-8")

def _read_xlsx(p):
    from openpyxl import load_workbook
    wb = load_workbook(p, read_only=True, data_only=True)
    sheets = {ws.title: [[c.value for c in row] for row in ws.iter_rows()] for ws in wb.worksheets}
    wb.close()
    return next(iter(sheets.values())) if len(sheets) == 1 else sheets
def _write_xlsx(p, data):
    from openpyxl import Workbook
    wb = Workbook()
    if isinstance(data, list):
        ws = wb.active
        for row in data:
            ws.append(row)
    else:
        for i, (name, rows) in enumerate(data.items()):
            ws = wb.active if i == 0 else wb.create_sheet(title=name)
            for row in rows:
                ws.append(row)
    wb.save(p)

def _read_sqlite(p):
    con = sqlite3.connect(p)
    con.row_factory = sqlite3.Row
    tables = {r["name"] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    out = {}
    for t in tables:
        out[t] = [dict(r) for r in con.execute(f'SELECT * FROM "{t}"')]
    con.close()
    return out
def _write_sqlite(p, data):
    con = sqlite3.connect(p)
    for table, rows in data.items():
        if not rows:
            continue
        cols = list(rows[0].keys())
        col_sql = ", ".join(f'"{c}"' for c in cols)
        ph = ", ".join("?" for _ in cols)
        con.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({col_sql})')
        for r in rows:
            con.execute(f'INSERT INTO "{table}" ({col_sql}) VALUES ({ph})',
                        tuple(r.get(c) for c in cols))
    con.commit()
    con.close()

def _read_text(p): return p.read_text(encoding="utf-8")
def _write_text(p, data): p.write_text(str(data), encoding="utf-8")


_READERS = {
    "json": _read_json, "yaml": _read_yaml, "yml": _read_yaml, "toml": _read_toml,
    "csv": _read_csv, "ini": _read_ini, "env": _read_env,
    "xlsx": _read_xlsx, "sqlite": _read_sqlite, "db": _read_sqlite,
}
_WRITERS = {
    "json": _write_json, "yaml": _write_yaml, "yml": _write_yaml, "toml": _write_toml,
    "csv": _write_csv, "ini": _write_ini, "env": _write_env,
    "xlsx": _write_xlsx, "sqlite": _write_sqlite, "db": _write_sqlite,
}


async def read_doc(path) -> Any:
    """按扩展名读取文件；未知格式按文本。"""
    p = Path(path)
    return _READERS.get(_ext(p), _read_text)(p)


async def write_doc(path, data) -> None:
    """按扩展名写回文件；未知格式按文本。"""
    p = Path(path)
    _WRITERS.get(_ext(p), _write_text)(p, data)


async def list_dir(path=".") -> list[dict]:
    p = Path(path)
    out = []
    for child in p.iterdir():
        out.append({
            "name": child.name,
            "path": str(child),
            "is_file": child.is_file(),
            "is_dir": child.is_dir(),
            "size": child.stat().st_size if child.is_file() else 0,
        })
    return sorted(out, key=lambda e: (not e["is_dir"], e["name"]))


async def stat_path(path) -> dict:
    p = Path(path)
    st = p.stat()
    return {
        "path": str(p),
        "is_file": p.is_file(),
        "is_dir": p.is_dir(),
        "size": st.st_size,
        "mtime": st.st_mtime,
    }
